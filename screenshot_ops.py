"""Generic HuggingFace causal-LM operator sequence capture via TorchDispatchMode.

Intercepts every aten op fired during a single meta-tensor forward pass and
writes the ordered sequence—with shapes, dtypes, module paths, and component
labels—to a formatted Excel workbook.

Supported model sources
-----------------------
* HF Hub model ID:  ``deepseek-ai/DeepSeek-V3``, ``Qwen/Qwen3-8B``, …
* Local directory:  ``./hf_models/deepseek_v3``,  ``/data/models/llama3``
  (must contain ``config.json``; custom architectures need an ``auto_map`` entry
  or ``configuration_*.py`` / ``modeling_*.py`` alongside the config)

Usage
-----
::

    python screenshot_ops.py <model_id_or_path> [options]
    python screenshot_ops.py deepseek-ai/DeepSeek-V3-0324 --layers 4
    python screenshot_ops.py Qwen/Qwen3-8B --layers 4 --seq-len 256
    python screenshot_ops.py ./hf_models/deepseek_v3 --layers 2
"""
from __future__ import annotations

import re
import sys
import json
import logging
import argparse
from pathlib import Path
from collections import Counter, defaultdict
from typing import Any, Dict, List, Optional, Tuple

import torch
import torch.nn as nn
from torch.utils._python_dispatch import TorchDispatchMode
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


# ── Compatibility shims ────────────────────────────────────────────────────────

def _apply_compat_patches() -> None:
    """Add deprecated transformers attributes expected by older model code.

    These are no-ops when the installed transformers version already exposes
    the attributes.  They must be called before any model module is imported
    so that the ``from transformers.xxx import yyy`` statements in model files
    succeed even on transformers >= 4.44.

    Also patches ``torch.autocast`` to accept ``device_type='meta'``.
    transformers 4.50+ passes the tensor's device type directly to autocast
    inside the RoPE implementation; meta tensors surface as ``'meta'``, which
    torch 2.x does not support.  We remap unknown device types to ``'cpu'``
    (autocast is a no-op for meta tensors regardless).
    """
    try:
        import transformers.utils.import_utils as _iu
        if not hasattr(_iu, "is_torch_fx_available"):
            _iu.is_torch_fx_available = lambda: True
    except ImportError:
        pass
    try:
        import transformers.utils as _tu
        if not hasattr(_tu, "is_torch_fx_available"):
            _tu.is_torch_fx_available = lambda: True
    except ImportError:
        pass
    try:
        import transformers.pytorch_utils as _pu
        if not hasattr(_pu, "is_torch_greater_or_equal_than_1_13"):
            _pu.is_torch_greater_or_equal_than_1_13 = True
    except ImportError:
        pass

    # Patch torch.autocast to handle unsupported device types (e.g. 'meta')
    _KNOWN_AUTOCAST_DEVICES = frozenset({"cpu", "cuda", "xpu", "hpu", "mps", "xla"})
    if not getattr(torch.amp.autocast, "_meta_device_safe", False):
        _orig_init = torch.amp.autocast.__init__

        def _safe_init(self, device_type: str, *args: Any, **kwargs: Any) -> None:
            if device_type not in _KNOWN_AUTOCAST_DEVICES:
                device_type = "cpu"
            _orig_init(self, device_type, *args, **kwargs)

        torch.amp.autocast.__init__ = _safe_init  # type: ignore[method-assign]
        torch.amp.autocast._meta_device_safe = True  # type: ignore[attr-defined]


# ── Zero-cost ops to skip ──────────────────────────────────────────────────────

_SKIP_OPS: frozenset = frozenset({
    "aten.detach.default", "aten.alias.default", "aten.t.default",
    "aten.view.default", "aten._unsafe_view.default",
    "aten.expand.default", "aten.contiguous.default",
    "aten.slice.Tensor", "aten.select.int",
    "aten.unsqueeze.default", "aten.squeeze.dim",
    "aten.split.Tensor", "aten.split_with_sizes.default",
    "aten.permute.default", "aten.reshape.default",
    "aten.clone.default",
    "aten.arange.default", "aten.arange.start",
    "aten.ones.default", "aten.zeros.default",
    "aten.full.default", "aten.scalar_tensor.default",
    "aten.tril.default", "aten.triu.default",
    "aten.empty_like.default", "aten.zeros_like.default",
    "aten.index_put_.default", "aten.index_put.default",
    "aten.scatter_.value", "aten.scatter_.src",
    "aten.histc.default", "aten.cumsum.default",
    "aten.bitwise_not.default",
    "aten.sort.default", "aten.sort.stable",
})


# ── Tensor helpers ─────────────────────────────────────────────────────────────

def _shape_str(t: torch.Tensor) -> str:
    return str(list(t.shape))


def _collect_tensors(args: tuple, kwargs: dict) -> List[torch.Tensor]:
    out: List[torch.Tensor] = []

    def _visit(x: Any) -> None:
        if isinstance(x, torch.Tensor):
            out.append(x)
        elif isinstance(x, (list, tuple)):
            for item in x:
                _visit(item)

    for a in args:
        _visit(a)
    for v in kwargs.values():
        _visit(v)
    return out


def _collect_output_tensors(out: Any) -> List[torch.Tensor]:
    if isinstance(out, torch.Tensor):
        return [out]
    if isinstance(out, (tuple, list)):
        return [x for x in out if isinstance(x, torch.Tensor)]
    return []


# ── Module tracker ─────────────────────────────────────────────────────────────

class _ModuleTracker:
    """Record which nn.Module is currently executing via lightweight hooks."""

    def __init__(self, root: nn.Module) -> None:
        self._stack: List[str] = []
        self._handles: List[Any] = []
        self._install(root, "")

    def _install(self, module: nn.Module, prefix: str) -> None:
        for name, child in module.named_children():
            full = f"{prefix}.{name}" if prefix else name

            def _pre(m: nn.Module, inp: Any, _fn: str = full) -> None:
                self._stack.append(_fn)

            def _post(m: nn.Module, inp: Any, out: Any, _fn: str = full) -> None:
                if self._stack and self._stack[-1] == _fn:
                    self._stack.pop()

            self._handles.append(child.register_forward_pre_hook(_pre))
            self._handles.append(child.register_forward_hook(_post))
            self._install(child, full)

    @property
    def current_module(self) -> str:
        return self._stack[-1] if self._stack else ""

    def remove(self) -> None:
        for h in self._handles:
            h.remove()
        self._handles.clear()


# ── Recording dispatch ─────────────────────────────────────────────────────────

class _RecordingDispatch(TorchDispatchMode):
    """Intercept every aten op and record its metadata."""

    def __init__(
        self,
        module_tracker: Optional[_ModuleTracker] = None,
        skip_reshapes: bool = True,
    ) -> None:
        super().__init__()
        self.records: List[Dict[str, Any]] = []
        self._tracker = module_tracker
        self._skip = skip_reshapes

    def __torch_dispatch__(self, func, types, args=(), kwargs=None):
        kwargs = kwargs or {}
        out = func(*args, **kwargs)

        op_name = f"{func.overloadpacket}.{func._overloadname}"
        if self._skip and op_name in _SKIP_OPS:
            return out

        in_t = _collect_tensors(args, kwargs)
        out_t = _collect_output_tensors(out)
        mod = self._tracker.current_module if self._tracker else ""

        self.records.append({
            "idx":           len(self.records),
            "aten_op":       op_name,
            "module_path":   mod,
            "layer":         _extract_layer_idx(mod),
            "component":     _classify_component(mod, op_name),
            "input_shapes":  ", ".join(_shape_str(t) for t in in_t),
            "input_dtypes":  ", ".join(str(t.dtype) for t in in_t),
            "output_shapes": ", ".join(_shape_str(t) for t in out_t),
            "output_dtypes": ", ".join(str(t.dtype) for t in out_t),
            "num_inputs":    len(in_t),
            "num_outputs":   len(out_t),
        })
        return out


# ── Component classification ───────────────────────────────────────────────────

def _extract_layer_idx(module_path: str) -> str:
    """Return the transformer-block index from a dotted module path, if present."""
    parts = module_path.split(".")
    # Common container names across architectures
    block_containers = {"layers", "blocks", "h", "layer"}
    for i, part in enumerate(parts):
        if part in block_containers and i + 1 < len(parts):
            try:
                return str(int(parts[i + 1]))
            except ValueError:
                pass
    return ""


def _classify_component(module_path: str, op_name: str) -> str:
    """Map a (module_path, aten_op) pair to a human-readable component label.

    Rules are ordered from specific to general and rely only on naming
    conventions shared across LLaMA, Qwen, Mistral, Mixtral, DeepSeek, etc.
    No architecture-specific class names are referenced.
    """
    s = module_path.lower()
    # Short op name, e.g. "mm" from "aten.mm.default"
    op = op_name.split(".")[1] if op_name.count(".") >= 1 else op_name

    # ── Norm layers ─────────────────────────────────────────────────────────
    _is_norm = any(tok in s for tok in (
        "layernorm", "rmsnorm", "layer_norm", "group_norm", "rms_norm"))
    if _is_norm:
        if "input" in s or "pre" in s:
            return "norm.pre_attn"
        if "post_attention" in s or "post_attn" in s:
            return "norm.post_attn"
        # Norm module that sits inside an attention block (e.g. q_a_layernorm)
        if "attn" in s or "attention" in s:
            return "attn.inner_norm"
        return "norm.final"

    # ── Attention ────────────────────────────────────────────────────────────
    _in_attn = "self_attn" in s or "attention" in s or (
        "attn" in s and "layernorm" not in s and "rmsnorm" not in s)

    if _in_attn:
        # Low-rank / LoRA-style projections (MLA in DeepSeek, etc.)
        for lora in ("q_a_proj", "q_b_proj", "kv_a_proj", "kv_b_proj"):
            if lora in s:
                return f"attn.{lora}"
        # Standard QKV and output projections
        for proj in ("q_proj", "k_proj", "v_proj", "o_proj",
                     "out_proj", "c_attn", "c_proj", "query_key_value"):
            if proj in s:
                return f"attn.{proj}"
        # Position-encoding helper
        if "rotary" in s or "rope" in s or "embed_positions" in s:
            return "attn.rope"
        # Core attention compute
        if op in ("matmul", "mm", "bmm"):
            return "attn.qk_score"
        if "softmax" in op or "safe_softmax" in op:
            return "attn.softmax"
        return f"attn.{op}"

    # ── MoE layers ───────────────────────────────────────────────────────────
    # Detect a bare 'gate' module (MoE router) vs 'gate_proj' (FFN projection).
    # e.g. "model.layers.3.mlp.gate"  → MoE router   (gate is a Module child)
    #      "model.layers.3.mlp.gate_proj" → FFN proj  (gate_proj is a Linear)
    _path_parts = s.split(".")
    _has_bare_gate = (
        "gate" in _path_parts
        and "gate_proj" not in s
        and "gate_up" not in s
    )
    _in_moe = (
        "expert" in s or "moe" in s or "mixture" in s or "sparse" in s
        or _has_bare_gate
    )

    if _in_moe:
        # Shared / always-on expert branch
        if "shared" in s:
            for proj in ("gate_proj", "up_proj", "down_proj",
                         "fc1", "fc2", "w1", "w2", "w3"):
                if proj in s:
                    return f"moe.shared.{proj}"
            return f"moe.shared.{op}"
        # Router / gate (no MLP projection suffix means this is the gating module)
        if "gate" in s and not any(p in s for p in (
                "gate_proj", "gate_up", "up_proj")):
            return f"moe.router.{op}"
        # Per-expert MLP projections
        for proj in ("gate_proj", "up_proj", "down_proj",
                     "fc1", "fc2", "w1", "w2", "w3"):
            if proj in s:
                return f"moe.expert.{proj}"
        return f"moe.expert.{op}"

    # ── Dense FFN / MLP ──────────────────────────────────────────────────────
    _in_ffn = "mlp" in s or "ffn" in s or "feed_forward" in s

    if _in_ffn:
        for proj in ("gate_proj", "up_proj", "down_proj",
                     "fc1", "fc2", "w1", "w2", "w3"):
            if proj in s:
                return f"ffn.{proj}"
        if op in ("silu", "gelu", "relu", "sigmoid", "gelu_new"):
            return "ffn.act"
        if op == "mul":
            return "ffn.mul"
        return f"ffn.{op}"

    # ── Embedding / LM head ──────────────────────────────────────────────────
    if "embed" in s:
        return "embedding"
    if "lm_head" in s or s.split(".")[-1] == "output":
        return "lm_head"
    # Final norm that sits outside any block
    if "norm" in s:
        return "norm.final"

    return op  # catch-all: bare aten op short name


# ── Generic MoE patch ──────────────────────────────────────────────────────────

def _is_moe_module(module: nn.Module) -> bool:
    """Duck-type check: does this module have a non-trivial ModuleList of experts?"""
    experts = getattr(module, "experts", None)
    return (
        isinstance(experts, nn.ModuleList)
        and any(e is not None for e in experts)
        and not getattr(module, "_meta_patched", False)
    )


def _patch_moe_for_meta(model: nn.Module) -> None:
    """Replace forward on MoE modules so they work with meta tensors.

    Real MoE forwards typically call ``.cpu().numpy()`` inside the routing
    kernel, which fails on meta tensors.  We detect affected modules by
    duck-typing (``experts`` is an ``nn.ModuleList``) and replace their
    forward with a simplified version that still fires every *type* of op
    (gate linear, expert MLP projections, shared-expert branch) while
    skipping the data-dependent dispatch loop.
    """
    patched = 0
    for _name, module in model.named_modules():
        if not _is_moe_module(module):
            continue
        module._meta_patched = True
        module.forward = _make_meta_moe_forward(module)
        patched += 1
        logger.debug("Patched MoE: %s (%d experts)",
                     type(module).__name__, len(module.experts))
    if patched:
        logger.info("Applied meta-tensor MoE patch to %d module(s).", patched)


def _returns_router_tuple(mod: nn.Module) -> bool:
    """Return True if this MoE module's forward yields (hidden, router_logits).

    Some architectures (Mixtral) return a 2-tuple; others (DeepSeek, Qwen-MoE)
    return a single tensor.  We detect this by inspecting the source code of
    the original forward method, with a duck-typing fallback.
    """
    import inspect
    try:
        src = inspect.getsource(type(mod).forward)
        # Common source-level markers for tuple-returning MoE forwards
        if any(pat in src for pat in ("router_logits", "aux_loss",
                                      "return hidden_states,")):
            return True
    except Exception:
        pass
    # Fallback: Mixtral-style uses 'router' sub-module; DeepSeek/Qwen use 'gate'
    return hasattr(mod, "router") and not hasattr(mod, "gate")


def _make_meta_moe_forward(mod: nn.Module):
    """Return a simplified forward closure bound to *mod*."""

    _tuple_return = _returns_router_tuple(mod)

    def _forward(hidden_states: torch.Tensor, *args: Any, **kwargs: Any):
        try:
            result = _impl(hidden_states)
            # Some callers unpack (hidden, router_logits) — return matching shape
            return (result, None) if _tuple_return else result
        except Exception as exc:
            logger.debug("Meta MoE forward error (%s) — returning identity.", exc)
            return (hidden_states, None) if _tuple_return else hidden_states

    def _impl(hidden_states: torch.Tensor) -> torch.Tensor:
        orig = hidden_states                         # (B, S, H)
        bs, seq, h = orig.shape
        flat = orig.reshape(bs * seq, h)             # (B*S, H)

        # ── Gate / router ────────────────────────────────────────────────────
        # Captures linear + activation + topk ops from the routing module.
        gate_weight: Optional[torch.Tensor] = None
        gate = getattr(mod, "gate", None)
        if gate is not None and callable(gate):
            try:
                # Try 3D first (e.g. DeepSeek MoEGate expects (B,S,H))
                gate_out = gate(orig)
                if isinstance(gate_out, (tuple, list)):
                    # (topk_idx, topk_weight) style — take weight only
                    gate_weight = gate_out[1]                 # (B*S, top_k)
                else:
                    gate_weight = torch.softmax(gate_out.float(), dim=-1)[:, :1]
            except Exception:
                try:
                    # Fallback: 2D input (B*S, H)
                    gate_out = gate(flat)
                    if isinstance(gate_out, (tuple, list)):
                        gate_weight = gate_out[1]
                    else:
                        gate_weight = torch.softmax(gate_out.float(), dim=-1)[:, :1]
                except Exception as exc:
                    logger.debug("Gate forward failed (%s), skipping weight.", exc)

        # ── One representative expert ─────────────────────────────────────────
        # Captures the expert MLP ops (projections + activations).
        first_expert = next((e for e in mod.experts if e is not None), None)
        if first_expert is None:
            return orig

        try:
            expert_out = first_expert(flat)              # (B*S, H)
        except Exception:
            # Some experts expect 3D; try that
            expert_out = first_expert(orig).reshape(bs * seq, -1)

        # ── Weighted combination ──────────────────────────────────────────────
        if gate_weight is not None:
            w = gate_weight[:, :1]  # (B*S, 1)
            y = expert_out * w
        else:
            y = expert_out

        # Reshape back to (B, S, H) — use -1 to handle any hidden-dim value
        try:
            y = y.reshape(bs, seq, -1)
        except Exception:
            y = orig

        # ── Shared experts ────────────────────────────────────────────────────
        # Captures the always-on shared expert branch (DeepSeek, etc.).
        for attr in ("shared_experts", "shared_expert"):
            shared = getattr(mod, attr, None)
            if shared is not None and callable(shared):
                try:
                    y = y + shared(orig)
                except Exception as exc:
                    logger.debug("Shared expert failed (%s), skipping.", exc)
                break

        return y

    return _forward


# ── Config normalization ───────────────────────────────────────────────────────

def _normalize_config(config: Any) -> None:
    """Apply compatibility fixes to a PretrainedConfig in-place.

    * ``rope_scaling``: older modeling code reads ``rope_scaling["type"]``
      but newer transformers saves the field as ``"rope_type"``.
    * ``_attn_implementation``: force ``"eager"`` so no flash-attn or SDPA
      kernel is needed during meta-tensor tracing.
    """
    rs = getattr(config, "rope_scaling", None)
    if isinstance(rs, dict) and "rope_type" in rs and "type" not in rs:
        rs["type"] = rs["rope_type"]

    config._attn_implementation = "eager"


# ── Model loading ──────────────────────────────────────────────────────────────

def load_model(
    model_id: str,
    num_hidden_layers: int = 4,
) -> Tuple[nn.Module, Any]:
    """Load any HF causal LM onto meta device for op-sequence tracing.

    Parameters
    ----------
    model_id:
        HuggingFace Hub ID (``"deepseek-ai/DeepSeek-V3"``) **or** a local
        directory path (``"./hf_models/deepseek_v3"``).
        Custom architectures must expose an ``auto_map`` entry in their
        ``config.json`` *or* be registered in the transformers model registry.
    num_hidden_layers:
        Number of transformer blocks to instantiate.  2–4 layers is enough
        to capture every distinct op pattern including both dense and MoE
        blocks (if applicable).

    Returns
    -------
    model, config
        The model is on meta device, in eval mode, with any MoE patch applied.
        ``config._full_num_hidden_layers`` preserves the original layer count.
    """
    from transformers import AutoConfig, AutoModelForCausalLM

    # Compat shims must run before any model module is imported
    _apply_compat_patches()

    logger.info("Loading config from %s …", model_id)
    config = AutoConfig.from_pretrained(
        model_id,
        trust_remote_code=True,
    )

    # Stash original depth before overriding
    config._full_num_hidden_layers = getattr(config, "num_hidden_layers", None)
    config.num_hidden_layers = num_hidden_layers
    _normalize_config(config)

    logger.info(
        "Instantiating %s on meta device (%d layers) …",
        type(config).__name__, num_hidden_layers,
    )
    with torch.device("meta"):
        model = AutoModelForCausalLM.from_config(config, trust_remote_code=True)
    model.eval()

    _patch_moe_for_meta(model)
    return model, config


# ── Config summary ─────────────────────────────────────────────────────────────

# Fields to include in the "Model Config" sheet when present on the config object.
# Order determines row order; architecture-specific fields appear after universal ones.
_SUMMARY_FIELDS: Tuple[str, ...] = (
    # ── Universal ──────────────────────────────────────────────────────────
    "model_type", "hidden_size", "intermediate_size",
    "num_hidden_layers", "num_attention_heads", "num_key_value_heads",
    "vocab_size", "max_position_embeddings", "rope_theta", "hidden_act",
    # ── MLA (DeepSeek MLA) ─────────────────────────────────────────────────
    "q_lora_rank", "kv_lora_rank",
    "qk_nope_head_dim", "qk_rope_head_dim", "v_head_dim",
    # ── MoE (universal field names) ────────────────────────────────────────
    "n_routed_experts",      # DeepSeek
    "num_local_experts",     # Mixtral
    "n_shared_experts",      # DeepSeek
    "num_experts_per_tok",   # DeepSeek / Qwen-MoE
    "num_experts_per_token", # alternative spelling
    "first_k_dense_replace", # DeepSeek
    "moe_intermediate_size", # DeepSeek
    "moe_layer_freq",        # DeepSeek
    # ── Misc ───────────────────────────────────────────────────────────────
    "attention_dropout", "sliding_window",
)


def build_config_summary(
    model_id: str,
    config: Any,
    num_traced: int,
    batch_size: int,
    seq_len: int,
) -> Dict[str, Any]:
    """Build an ordered dict of config values for the Excel summary sheet."""
    summary: Dict[str, Any] = {
        "model_id":               model_id,
        "full_num_hidden_layers": getattr(config, "_full_num_hidden_layers", "?"),
        "traced_num_hidden_layers": num_traced,
        "batch_size":             batch_size,
        "seq_len":                seq_len,
    }
    for field in _SUMMARY_FIELDS:
        val = getattr(config, field, None)
        if val is not None:
            summary[field] = val
    return summary


# ── Excel writing ──────────────────────────────────────────────────────────────

# Component-prefix → background colour (generic, architecture-agnostic)
_FILL_COLORS: Dict[str, str] = {
    "norm.":       "E8F5E9",   # light green  — all norm layers
    "attn.":       "E3F2FD",   # light blue   — attention
    "moe.router.": "FFF3E0",   # light orange — MoE router / gate
    "moe.shared.": "FFF8E1",   # light yellow — always-on shared expert
    "moe.expert.": "FCE4EC",   # light pink   — routed experts
    "ffn.":        "F3E5F5",   # light purple — dense FFN
    "embedding":   "ECEFF1",   # light grey   — embedding table
    "lm_head":     "ECEFF1",   # light grey   — language-model head
}


def _get_fill(component: str) -> Optional[PatternFill]:
    for prefix, color in _FILL_COLORS.items():
        if component.startswith(prefix):
            return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return None


def write_excel(
    records: List[Dict[str, Any]],
    output_path: Path,
    config_summary: Dict[str, Any],
) -> None:
    """Write op records to a multi-sheet formatted Excel workbook."""
    wb = openpyxl.Workbook()
    bold12 = Font(bold=True, size=12)
    hdr_fill = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style="thin", color="BDBDBD"))

    # ── Sheet 1: Model Config ───────────────────────────────────────────────
    ws_cfg = wb.active
    ws_cfg.title = "Model Config"
    ws_cfg.append(["Parameter", "Value"])
    ws_cfg["A1"].font = bold12
    ws_cfg["B1"].font = bold12
    for key, val in config_summary.items():
        ws_cfg.append([key, str(val)])
    ws_cfg.column_dimensions["A"].width = 32
    ws_cfg.column_dimensions["B"].width = 45

    # ── Sheet 2: Operator Sequence ──────────────────────────────────────────
    ws = wb.create_sheet("Operator Sequence")
    columns = [
        ("Index",         8),
        ("Aten Op",      36),
        ("Module Path",  55),
        ("Layer",         7),
        ("Component",    26),
        ("Input Shapes", 50),
        ("Input Dtypes", 30),
        ("Output Shapes",50),
        ("Output Dtypes",30),
    ]
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, rec in enumerate(records, 2):
        values = [
            rec["idx"], rec["aten_op"], rec["module_path"], rec["layer"],
            rec["component"], rec["input_shapes"], rec["input_dtypes"],
            rec["output_shapes"], rec["output_dtypes"],
        ]
        fill = _get_fill(rec["component"])
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"
    ws.freeze_panes = "A2"

    # ── Sheet 3: Summary by Component ──────────────────────────────────────
    ws_comp = wb.create_sheet("By Component")
    ws_comp.append(["Component", "Count"])
    ws_comp["A1"].font = bold12
    ws_comp["B1"].font = bold12
    for comp, cnt in sorted(Counter(r["component"] for r in records).items()):
        ws_comp.append([comp, cnt])
    ws_comp.column_dimensions["A"].width = 30
    ws_comp.column_dimensions["B"].width = 10

    # ── Sheet 4: Summary by Layer ───────────────────────────────────────────
    ws_layer = wb.create_sheet("By Layer")
    ws_layer.append(["Layer", "Op Count", "Components"])
    for cell in ws_layer["1:1"]:
        cell.font = bold12
    layer_info: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"count": 0, "components": set()})
    for r in records:
        lk = r["layer"] or "non-layer"
        layer_info[lk]["count"] += 1
        layer_info[lk]["components"].add(r["component"])

    for lk in sorted(layer_info, key=lambda x: (x == "non-layer", x)):
        info = layer_info[lk]
        ws_layer.append([lk, info["count"], ", ".join(sorted(info["components"]))])
    ws_layer.column_dimensions["A"].width = 12
    ws_layer.column_dimensions["B"].width = 10
    ws_layer.column_dimensions["C"].width = 80

    wb.save(output_path)
    logger.info("Saved %d ops → %s", len(records), output_path)


# ── High-level pipeline ────────────────────────────────────────────────────────

def run_trace(
    model_id: str,
    num_layers: int = 4,
    batch_size: int = 1,
    seq_len: int = 128,
    output_path: Optional[Path] = None,
) -> Tuple[Path, List[Dict[str, Any]]]:
    """Full pipeline: load → trace → write Excel.

    Parameters
    ----------
    model_id:
        HF Hub ID or local path.
    num_layers:
        Number of transformer blocks to trace.
    batch_size, seq_len:
        Shape of the dummy input.
    output_path:
        Destination ``.xlsx`` file.  Defaults to ``<model_slug>_ops.xlsx``
        in the current directory.

    Returns
    -------
    (output_path, records)
        ``records`` is the raw list of op dicts, useful for programmatic
        inspection in tests.
    """
    model, config = load_model(model_id, num_layers)

    vocab = getattr(config, "vocab_size", 32000)
    input_ids   = torch.randint(0, vocab, (batch_size, seq_len), device="meta")
    position_ids = torch.arange(seq_len, device="meta").unsqueeze(0)
    mask = torch.triu(
        torch.full((1, 1, seq_len, seq_len), float("-inf"), device="meta"),
        diagonal=1,
    )

    logger.info(
        "Tracing forward pass  model=%s  batch=%d  seq=%d …",
        model_id, batch_size, seq_len,
    )
    tracker  = _ModuleTracker(model)
    recorder = _RecordingDispatch(module_tracker=tracker, skip_reshapes=True)
    try:
        with recorder, torch.no_grad():
            try:
                model(input_ids=input_ids, attention_mask=mask,
                      position_ids=position_ids, use_cache=False)
            except Exception as e_full:
                # Some models ignore unknown kwargs; retry with minimal args
                logger.warning(
                    "Full forward failed (%s), retrying with input_ids only …", e_full)
                recorder.records.clear()
                model(input_ids=input_ids, use_cache=False)
    finally:
        tracker.remove()

    logger.info(
        "Captured %d ops (after filtering zero-cost reshapes).",
        len(recorder.records),
    )

    if output_path is None:
        slug = re.sub(r"[^\w-]", "_", model_id).strip("_")
        output_path = Path(f"{slug}_ops.xlsx")

    summary = build_config_summary(
        model_id, config, num_layers, batch_size, seq_len)
    write_excel(recorder.records, output_path, summary)

    return output_path, recorder.records


# ── CLI entry point ────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Capture the aten op sequence of a HuggingFace causal LM.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "model_id",
        help="HF Hub model ID or local path  "
             "(e.g. 'deepseek-ai/DeepSeek-V3-0324', 'Qwen/Qwen3-8B', "
             "'./hf_models/deepseek_v3')",
    )
    parser.add_argument(
        "--layers", "-l", type=int, default=4, metavar="N",
        help="Number of transformer blocks to trace",
    )
    parser.add_argument(
        "--output", "-o", type=Path, default=None, metavar="FILE",
        help="Output .xlsx path (default: <model_slug>_ops.xlsx)",
    )
    parser.add_argument(
        "--batch-size", "-b", type=int, default=1, metavar="B",
    )
    parser.add_argument(
        "--seq-len", "-s", type=int, default=128, metavar="S",
    )
    args = parser.parse_args()

    out, records = run_trace(
        model_id=args.model_id,
        num_layers=args.layers,
        batch_size=args.batch_size,
        seq_len=args.seq_len,
        output_path=args.output,
    )
    print(f"\nWritten {len(records)} ops → {out}")


if __name__ == "__main__":
    main()
